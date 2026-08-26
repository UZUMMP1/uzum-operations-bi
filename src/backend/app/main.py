from collections import defaultdict
from datetime import date, timedelta
import io
from math import ceil
import os
from pathlib import Path
import re
from urllib.parse import quote
from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, func, inspect as sa_inspect, select
from sqlalchemy.orm import Session
from .database import Base, SessionLocal, engine
from .importers import import_file, parse_seller_sku_color
from .metrics import calculate_metrics, safe_divide
from .models import InventoryHistory, ProductMapping, SalesRecord, TrafficHistory, UploadBatch

Base.metadata.create_all(bind=engine)
with engine.begin() as connection:
    inventory_columns = {column["name"] for column in sa_inspect(connection).get_columns("inventory_history")}
    if "seller_sku" not in inventory_columns:
        connection.exec_driver_sql("ALTER TABLE inventory_history ADD COLUMN seller_sku VARCHAR(255)")
app = FastAPI(title="UZUM BI API", version="0.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"], allow_methods=["*"], allow_headers=["*"])
PUBLIC_READ_ONLY = os.getenv("UZUM_PUBLIC_READ_ONLY") == "1"

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "parser_version": "uzum-samples-v4-shifted-inventory-date",
        "read_only": PUBLIC_READ_ONLY,
        "database": engine.dialect.name,
    }

def _upload_batch_dict(batch: UploadBatch) -> dict:
    return {
        "id": batch.id,
        "data_type": batch.data_type,
        "filename": batch.filename,
        "uploaded_at": batch.uploaded_at,
        "row_count": batch.row_count,
        "status": batch.status,
        "error_message": batch.error_message,
    }

def _date_range(start: date, end: date) -> set[date]:
    return {start + timedelta(days=offset) for offset in range((end - start).days + 1)}

def _data_quality_overview(db: Session, start: date, end: date, category: str = "all") -> dict:
    if start > end:
        raise HTTPException(status_code=422, detail="开始日期不能晚于结束日期")

    category_skus = _category_skus(db, category)
    sales_dates_query = select(SalesRecord.created_date).where(SalesRecord.created_date.between(start, end)).distinct()
    traffic_dates_query = select(TrafficHistory.record_date).where(TrafficHistory.record_date.between(start, end)).distinct()
    sales_rows_query = select(func.count()).select_from(SalesRecord).where(SalesRecord.created_date.between(start, end))
    traffic_rows_query = select(func.count()).select_from(TrafficHistory).where(TrafficHistory.record_date.between(start, end))
    if category_skus is not None:
        sales_dates_query = sales_dates_query.where(SalesRecord.sku.in_(category_skus))
        traffic_dates_query = traffic_dates_query.where(TrafficHistory.sku.in_(category_skus))
        sales_rows_query = sales_rows_query.where(SalesRecord.sku.in_(category_skus))
        traffic_rows_query = traffic_rows_query.where(TrafficHistory.sku.in_(category_skus))

    sales_dates = set(db.scalars(sales_dates_query))
    traffic_dates = set(db.scalars(traffic_dates_query))
    expected_sales_dates = _date_range(start, end)
    missing_sales_dates = sorted(expected_sales_dates - sales_dates)
    missing_traffic_dates = sorted(sales_dates - traffic_dates)

    inventory_required = {start + timedelta(days=1), end + timedelta(days=1)}
    inventory_dates_query = select(InventoryHistory.download_date).where(
        InventoryHistory.download_date.in_(inventory_required)
    ).distinct()
    inventory_rows_query = select(func.count()).select_from(InventoryHistory).where(
        InventoryHistory.download_date.between(start + timedelta(days=1), end + timedelta(days=1))
    )
    if category_skus is not None:
        inventory_dates_query = inventory_dates_query.where(InventoryHistory.sku.in_(category_skus))
        inventory_rows_query = inventory_rows_query.where(InventoryHistory.sku.in_(category_skus))
    inventory_dates = set(db.scalars(inventory_dates_query))
    missing_inventory_dates = sorted(inventory_required - inventory_dates)

    uploads = list(db.scalars(select(UploadBatch).order_by(UploadBatch.uploaded_at.desc()).limit(12)))
    issues = []
    if not sales_dates:
        issues.append({"severity": "critical", "code": "sales_missing", "title": "所选周期没有销售数据", "detail": "经营指标、日报和 Agent 结论无法生成。", "action": "上传覆盖所选日期的销售底表。"})
    elif missing_sales_dates:
        issues.append({"severity": "warning", "code": "sales_gap", "title": f"销售数据缺少 {len(missing_sales_dates)} 个日期", "detail": "、".join(map(str, missing_sales_dates[:8])) + (" 等" if len(missing_sales_dates) > 8 else ""), "action": "确认这些日期是否确实无销售；否则重传累计销售底表。"})
    if missing_traffic_dates:
        issues.append({"severity": "warning", "code": "traffic_gap", "title": f"流量数据缺少 {len(missing_traffic_dates)} 个销售日期", "detail": "缺失日期的 UV/CVR 不参与计算。", "action": "补充对应日期的商品分析流量报表。"})
    if missing_inventory_dates:
        issues.append({"severity": "warning", "code": "inventory_gap", "title": "期初或期末库存快照缺失", "detail": f"需要原始库存 DATE={ '、'.join(map(str, missing_inventory_dates)) }。", "action": "上传包含对应 DATE 的累计库存底表。"})
    if not issues:
        issues.append({"severity": "info", "code": "complete", "title": "所选周期关键数据完整", "detail": "销售日期、流量覆盖和边界库存快照均可用。", "action": "可以生成日报、周报和运营 Agent 分析。"})

    return {
        "period": {"start": start, "end": end},
        "complete": not any(item["severity"] in ("critical", "warning") for item in issues),
        "sources": {
            "sales": {"label": "销售", "covered_days": len(sales_dates), "expected_days": len(expected_sales_dates), "row_count": int(db.scalar(sales_rows_query) or 0), "latest_date": max(sales_dates) if sales_dates else None, "missing_dates": missing_sales_dates[:14]},
            "traffic": {"label": "流量", "covered_days": len(sales_dates & traffic_dates), "expected_days": len(sales_dates), "row_count": int(db.scalar(traffic_rows_query) or 0), "latest_date": max(traffic_dates) if traffic_dates else None, "missing_dates": missing_traffic_dates[:14]},
            "inventory": {"label": "库存", "covered_days": len(inventory_required & inventory_dates), "expected_days": len(inventory_required), "row_count": int(db.scalar(inventory_rows_query) or 0), "latest_date": max(inventory_dates) if inventory_dates else None, "missing_dates": missing_inventory_dates},
        },
        "issues": issues,
        "uploads": [_upload_batch_dict(batch) for batch in uploads],
    }

@app.get("/api/data-quality/overview")
def data_quality_overview(start: date = Query(...), end: date = Query(...), category: str = "all", db: Session = Depends(get_db)):
    return _data_quality_overview(db, start, end, category)

@app.get("/api/uploads/history")
def upload_history(limit: int = Query(20, ge=1, le=100), db: Session = Depends(get_db)):
    batches = db.scalars(select(UploadBatch).order_by(UploadBatch.uploaded_at.desc()).limit(limit))
    return {"items": [_upload_batch_dict(batch) for batch in batches]}

@app.get("/api/dashboard/summary")
def dashboard_summary(start: date = Query(...), end: date = Query(...), category: str = "all", db: Session = Depends(get_db)):
    return _period_metrics(db, start, end, category)

@app.get("/api/dashboard/trend")
def dashboard_trend(start: date = Query(...), end: date = Query(...), category: str = "all", db: Session = Depends(get_db)):
    return _trend(db, start, end, "day", None, category)

@app.get("/api/sales/trend")
def sales_trend(start: date = Query(...), end: date = Query(...), granularity: str = Query("day", pattern="^(day|week|month)$"), sku: str | None = None, category: str = "all", db: Session = Depends(get_db)):
    return _trend(db, start, end, granularity, sku, category)

def _bucket(day: date, granularity: str) -> date:
    if granularity == "week":
        return day - timedelta(days=day.weekday())
    if granularity == "month":
        return day.replace(day=1)
    return day

def _trend(db: Session, start: date, end: date, granularity: str, sku: str | None, category: str = "all"):
    sales_query = select(SalesRecord.created_date, SalesRecord.quantity, SalesRecord.returns, SalesRecord.gmv).where(SalesRecord.created_date.between(start, end))
    traffic_query = select(TrafficHistory.record_date, TrafficHistory.uv).where(TrafficHistory.record_date.between(start, end))
    if sku:
        sales_query = sales_query.where(SalesRecord.sku == sku)
        traffic_query = traffic_query.where(TrafficHistory.sku == sku)
    category_skus = _category_skus(db, category)
    if category_skus is not None:
        sales_query = sales_query.where(SalesRecord.sku.in_(category_skus))
        traffic_query = traffic_query.where(TrafficHistory.sku.in_(category_skus))
    sales_rows = list(db.execute(sales_query))
    traffic_rows = list(db.execute(traffic_query))
    traffic_dates = {day for day, _ in traffic_rows}
    buckets = defaultdict(lambda: {"so": 0.0, "orders": 0.0, "returns": 0.0, "gmv": 0.0, "uv": None, "cvr_quantity": 0.0})
    for day, quantity, returns, gmv in sales_rows:
        item = buckets[_bucket(day, granularity)]
        item["so"] += float(quantity - returns); item["orders"] += float(quantity); item["returns"] += float(returns); item["gmv"] += float(gmv)
        if day in traffic_dates:
            item["cvr_quantity"] += float(quantity)
    for day, uv in traffic_rows:
        item = buckets[_bucket(day, granularity)]
        item["uv"] = (item["uv"] or 0) + float(uv or 0)
    result = []
    for day, item in sorted(buckets.items()):
        cvr_quantity = item.pop("cvr_quantity")
        result.append({"date": day, **item, "asp": safe_divide(item["gmv"], item["so"]), "cvr": safe_divide(cvr_quantity, item["uv"]) if item["uv"] is not None else None})
    return result

def _latest_inventory(db: Session, end: date) -> tuple[date | None, dict[str, dict]]:
    # UZUM's inventory DATE describes the next sales day boundary. Sales day D
    # therefore uses only the exact raw inventory DATE D+1; never backfill from
    # an earlier snapshot because that would silently invent stock data.
    source_snapshot = end + timedelta(days=1)
    exists = db.scalar(select(func.count()).select_from(InventoryHistory).where(InventoryHistory.download_date == source_snapshot))
    if not exists:
        return None, {}
    rows = db.execute(select(
        InventoryHistory.sku,
        func.max(InventoryHistory.product),
        func.max(InventoryHistory.seller_sku),
        func.max(InventoryHistory.color),
        func.max(InventoryHistory.memory),
        func.max(InventoryHistory.region),
        func.sum(InventoryHistory.inventory),
    ).where(InventoryHistory.download_date == source_snapshot).group_by(InventoryHistory.sku)).all()
    return source_snapshot, {row[0]: {"product": row[1], "seller_sku": row[2], "color": row[3] or parse_seller_sku_color(row[2] or ""), "memory": row[4], "region": row[5], "inventory": float(row[6] or 0)} for row in rows}

def _sku_catalog(db: Session) -> dict[str, dict]:
    """Return the newest known identity for every barcode, independent of report dates."""
    rows = db.execute(select(
        InventoryHistory.sku,
        InventoryHistory.product,
        InventoryHistory.seller_sku,
        InventoryHistory.color,
        InventoryHistory.memory,
        InventoryHistory.region,
    ).order_by(InventoryHistory.download_date.desc(), InventoryHistory.id.desc())).all()
    catalog: dict[str, dict] = {}
    for sku, product, seller_sku, color, memory, region in rows:
        if sku not in catalog:
            catalog[sku] = {"product": product, "seller_sku": seller_sku, "color": color or parse_seller_sku_color(seller_sku or ""), "memory": memory, "region": region}
    return catalog

def _business_category(item: dict) -> str:
    text = " ".join(str(item.get(key) or "") for key in ("category", "product", "seller_sku")).lower()
    if "планшет" in text or "tablet" in text:
        return "平板"
    if "смартфон" in text or "smartphone" in text or re.search(r"\bphone\b", text):
        return "手机"
    return "可穿戴及其他"

def _category_skus(db: Session, category: str | None) -> set[str] | None:
    """Resolve the global UI category to the complete barcode set."""
    if not category or category == "all":
        return None
    target = {"phone": "手机", "tablet": "平板", "aiot": "可穿戴及其他"}.get(category)
    if target is None:
        return None
    catalog = _sku_catalog(db)
    sales_categories = {
        sku: raw_category
        for sku, raw_category in db.execute(
            select(SalesRecord.sku, func.max(SalesRecord.category)).group_by(SalesRecord.sku)
        )
    }
    all_skus = set(catalog) | set(sales_categories)
    return {
        sku for sku in all_skus
        if _business_category({**catalog.get(sku, {}), "category": sales_categories.get(sku)}) == target
    }

def _category_breakdown(products: list[dict]) -> list[dict]:
    totals = {label: {"so": 0.0, "orders": 0.0, "gmv": 0.0, "uv": 0.0, "has_uv": False} for label in ("手机", "平板", "AIOT")}
    for item in products:
        raw_label = _business_category(item)
        label = raw_label if raw_label in ("手机", "平板") else "AIOT"
        totals[label]["so"] += item["so"]
        totals[label]["orders"] += item["orders"]
        totals[label]["gmv"] += item["gmv"]
        if item.get("uv") is not None:
            totals[label]["uv"] += item["uv"]
            totals[label]["has_uv"] = True
    total_so = sum(value["so"] for value in totals.values())
    total_gmv = sum(value["gmv"] for value in totals.values())
    return [{
        "category": key,
        "so": value["so"],
        "orders": value["orders"],
        "gmv": value["gmv"],
        "uv": value["uv"] if value["has_uv"] else None,
        "asp": safe_divide(value["gmv"], value["so"]),
        "cvr": safe_divide(value["orders"], value["uv"]) if value["has_uv"] else None,
        "so_share": safe_divide(value["so"], total_so),
        "gmv_share": safe_divide(value["gmv"], total_gmv),
    } for key, value in totals.items()]

def _sku_metrics(db: Session, start: date, end: date, category: str = "all") -> list[dict]:
    category_skus = _category_skus(db, category)
    sales_query = select(SalesRecord.sku, func.max(SalesRecord.category), func.sum(SalesRecord.quantity), func.sum(SalesRecord.returns), func.sum(SalesRecord.gmv)).where(SalesRecord.created_date.between(start, end))
    traffic_query = select(TrafficHistory.record_date, TrafficHistory.sku, TrafficHistory.uv).where(TrafficHistory.record_date.between(start, end))
    if category_skus is not None:
        sales_query = sales_query.where(SalesRecord.sku.in_(category_skus))
        traffic_query = traffic_query.where(TrafficHistory.sku.in_(category_skus))
    sales_rows = db.execute(sales_query.group_by(SalesRecord.sku)).all()
    traffic_detail = db.execute(traffic_query).all()
    traffic = defaultdict(float)
    for _, sku, uv in traffic_detail:
        traffic[sku] += float(uv or 0)
    inventory_snapshot, inventory = _latest_inventory(db, end)
    catalog = _sku_catalog(db)
    result = []
    for sku, category, quantity, returns, gmv in sales_rows:
        so = float(quantity or 0) - float(returns or 0); uv = traffic.get(sku)
        current_stock = inventory.get(sku, {})
        identity = current_stock or catalog.get(sku, {})
        result.append({"sku": sku, "product": identity.get("product"), "seller_sku": identity.get("seller_sku"), "color": identity.get("color"), "memory": identity.get("memory"), "region": identity.get("region"), "category": category, "so": so, "orders": float(quantity or 0), "returns": float(returns or 0), "gmv": float(gmv or 0), "asp": safe_divide(float(gmv or 0), so), "uv": uv, "cvr": safe_divide(float(quantity or 0), uv) if uv is not None else None, "return_rate": safe_divide(float(returns or 0), float(quantity or 0)), "inventory": current_stock.get("inventory", 0) if inventory_snapshot is not None else None})
    return result

@app.get("/api/sales/products")
def sales_products(start: date = Query(...), end: date = Query(...), search: str | None = None, category: str = "all", db: Session = Depends(get_db)):
    rows = _sku_metrics(db, start, end, category)
    days = (end - start).days + 1
    previous_end = start - timedelta(days=1)
    previous_start = previous_end - timedelta(days=days - 1)
    previous = {item["sku"]: item for item in _sku_metrics(db, previous_start, previous_end, category)}
    for row in rows:
        prior = previous.get(row["sku"], {})
        row["comparisons"] = {
            key: _change(row.get(key), prior.get(key, 0))
            for key in ("so", "orders", "gmv", "asp", "uv", "cvr")
        }
    if search:
        needle = search.lower(); rows = [row for row in rows if needle in row["sku"].lower() or needle in (row["product"] or "").lower()]
    return sorted(rows, key=lambda row: row["gmv"], reverse=True)

@app.get("/api/dashboard/categories")
def dashboard_categories(start: date = Query(...), end: date = Query(...), category: str = "all", db: Session = Depends(get_db)):
    categories = []
    for key, label in (("phone", "手机"), ("tablet", "平板"), ("aiot", "AIOT")):
        metrics = _period_metrics(db, start, end, key) if category in ("all", key) else {"so": 0, "orders": 0, "gmv": 0, "uv": None, "asp": None, "cvr": None}
        categories.append({"category": label, **{field: metrics.get(field) for field in ("so", "orders", "gmv", "uv", "asp", "cvr")}})
    total_so = sum(item["so"] for item in categories)
    total_orders = sum(item["orders"] for item in categories)
    total_gmv = sum(item["gmv"] for item in categories)
    known_uv = [item["uv"] for item in categories if item["uv"] is not None]
    total_uv = sum(known_uv) if known_uv else None
    for item in categories:
        item["so_share"] = safe_divide(item["so"], total_so)
        item["gmv_share"] = safe_divide(item["gmv"], total_gmv)
    return {"items": categories, "total": {"so": total_so, "orders": total_orders, "gmv": total_gmv, "uv": total_uv, "asp": safe_divide(total_gmv, total_so), "cvr": safe_divide(total_orders, total_uv) if total_uv is not None else None}}

def _change(current: float | None, previous: float | None) -> float | None:
    if current is None or previous in (None, 0):
        return None
    return (current - previous) / previous

def _daily_diagnosis(db: Session, day: date, category: str = "all") -> dict:
    """Deterministic day-over-day diagnosis used by the overview Agent card."""
    previous_day = day - timedelta(days=1)
    current = _period_metrics(db, day, day, category)
    previous = _period_metrics(db, previous_day, previous_day, category)
    changes = {
        key: _change(current.get(key), previous.get(key))
        for key in ("so", "orders", "gmv", "asp", "uv", "cvr", "return_rate")
    }

    current_products = {item["sku"]: item for item in _sku_metrics(db, day, day, category)}
    previous_products = {item["sku"]: item for item in _sku_metrics(db, previous_day, previous_day, category)}
    catalog = _sku_catalog(db)
    all_skus = set(current_products) | set(previous_products)
    sku_rows = []
    for sku in all_skus:
        current_item = current_products.get(sku, {})
        previous_item = previous_products.get(sku, {})
        identity = current_item or previous_item or catalog.get(sku, {})
        current_so = float(current_item.get("so", 0))
        previous_so = float(previous_item.get("so", 0))
        sku_rows.append({
            "sku": sku,
            "product": identity.get("product"),
            "seller_sku": identity.get("seller_sku"),
            "color": identity.get("color"),
            "memory": identity.get("memory"),
            "region": identity.get("region"),
            "current_so": current_so,
            "previous_so": previous_so,
            "delta_so": current_so - previous_so,
        })
    is_growth = current["so"] > previous["so"]
    sku_drivers = sorted(sku_rows, key=lambda item: item["delta_so"], reverse=is_growth)[:5]

    current_categories = {item["category"]: item for item in _category_breakdown(list(current_products.values()))}
    previous_categories = {item["category"]: item for item in _category_breakdown(list(previous_products.values()))}
    category_drivers = []
    for category in set(current_categories) | set(previous_categories):
        current_so = float(current_categories.get(category, {}).get("so", 0))
        previous_so = float(previous_categories.get(category, {}).get("so", 0))
        category_drivers.append({"category": category, "current_so": current_so, "previous_so": previous_so, "delta_so": current_so - previous_so})
    category_drivers.sort(key=lambda item: item["delta_so"], reverse=is_growth)

    current_source, current_inventory = _latest_inventory(db, day)
    previous_source, _ = _latest_inventory(db, previous_day)
    stockouts = []
    for sku, previous_item in previous_products.items():
        stock = current_inventory.get(sku)
        if stock is not None and stock["inventory"] <= 0 and previous_item["so"] > 0:
            identity = stock or catalog.get(sku, {})
            stockouts.append({
                "sku": sku,
                "product": identity.get("product"),
                "seller_sku": identity.get("seller_sku"),
                "color": identity.get("color"),
                "memory": identity.get("memory"),
                "region": identity.get("region"),
                "previous_so": previous_item["so"],
                "current_so": current_products.get(sku, {}).get("so", 0),
                "inventory": stock["inventory"],
            })
    stockouts.sort(key=lambda item: item["previous_so"], reverse=True)

    category_skus = _category_skus(db, category)
    sales_current_query = select(func.count()).select_from(SalesRecord).where(SalesRecord.created_date == day)
    sales_previous_query = select(func.count()).select_from(SalesRecord).where(SalesRecord.created_date == previous_day)
    traffic_current_query = select(func.count()).select_from(TrafficHistory).where(TrafficHistory.record_date == day)
    traffic_previous_query = select(func.count()).select_from(TrafficHistory).where(TrafficHistory.record_date == previous_day)
    if category_skus is not None:
        sales_current_query = sales_current_query.where(SalesRecord.sku.in_(category_skus))
        sales_previous_query = sales_previous_query.where(SalesRecord.sku.in_(category_skus))
        traffic_current_query = traffic_current_query.where(TrafficHistory.sku.in_(category_skus))
        traffic_previous_query = traffic_previous_query.where(TrafficHistory.sku.in_(category_skus))
    sales_current = bool(db.scalar(sales_current_query))
    sales_previous = bool(db.scalar(sales_previous_query))
    traffic_current = bool(db.scalar(traffic_current_query))
    traffic_previous = bool(db.scalar(traffic_previous_query))
    quality = {
        "sales_current": sales_current,
        "sales_previous": sales_previous,
        "traffic_current": traffic_current,
        "traffic_previous": traffic_previous,
        "inventory_current": current_source is not None,
        "inventory_previous": previous_source is not None,
    }
    quality["complete"] = all(quality.values())

    def metric_check(key: str, label: str, adverse_when_up: bool = False) -> dict:
        change = changes[key]
        if current.get(key) is None or previous.get(key) is None:
            status = "missing"
        elif change is None:
            status = "neutral"
        elif (change > 0) == adverse_when_up:
            status = "negative"
        elif change == 0:
            status = "neutral"
        else:
            status = "positive"
        return {"key": key, "label": label, "current": current.get(key), "previous": previous.get(key), "change": change, "status": status}

    checks = [
        metric_check("uv", "UV是否下降"),
        metric_check("cvr", "CVR是否下降"),
        metric_check("asp", "ASP是否变化"),
        {"key": "stockout", "label": "重点SKU是否缺货", "status": "missing" if current_source is None else ("negative" if stockouts else "positive"), "count": len(stockouts)},
        metric_check("return_rate", "取消/退款率是否上升", adverse_when_up=True),
        {"key": "sku", "label": "哪些SKU对上升贡献最大" if is_growth else "哪些SKU对下降贡献最大", "status": "positive" if is_growth and any(item["delta_so"] > 0 for item in sku_drivers) else ("negative" if any(item["delta_so"] < 0 for item in sku_drivers) else "neutral"), "count": sum(item["delta_so"] > 0 if is_growth else item["delta_so"] < 0 for item in sku_drivers)},
        {"key": "category", "label": "哪些品类拉动最大" if is_growth else "哪些品类拖累最大", "status": "positive" if is_growth and any(item["delta_so"] > 0 for item in category_drivers) else ("negative" if any(item["delta_so"] < 0 for item in category_drivers) else "neutral"), "count": sum(item["delta_so"] > 0 if is_growth else item["delta_so"] < 0 for item in category_drivers)},
        {"key": "quality", "label": "数据是否完整", "status": "positive" if quality["complete"] else "missing"},
    ]
    so_change = changes["so"]
    if not sales_current or not sales_previous:
        headline = "销售数据不完整，暂不能判断日对日销量变化原因"
        direction = "missing"
    elif so_change is None:
        headline = f"{day} 实际销量为 {current['so']:,.0f}，前一天无可比基数"
        direction = "flat"
    elif so_change < 0:
        headline = f"{day} 实际销量较前一天下降 {abs(so_change):.1%}"
        direction = "decline"
    elif so_change > 0:
        headline = f"{day} 实际销量较前一天增长 {so_change:.1%}"
        direction = "growth"
    else:
        headline = f"{day} 实际销量与前一天持平"
        direction = "flat"
    return {
        "date": day,
        "previous_date": previous_day,
        "direction": direction,
        "headline": headline,
        "current": current,
        "previous": previous,
        "changes": changes,
        "checks": checks,
        "stockouts": stockouts[:5],
        "sku_drivers": sku_drivers,
        "category_drivers": category_drivers,
        "data_quality": quality,
    }

@app.get("/api/dashboard/daily-diagnosis")
def daily_diagnosis(day: date = Query(...), category: str = "all", db: Session = Depends(get_db)):
    return _daily_diagnosis(db, day, category)

@app.get("/api/returns")
def returns_monitor(start: date = Query(...), end: date = Query(...), db: Session = Depends(get_db), category: str = "all"):
    returns_query = select(SalesRecord.created_date, SalesRecord.sku, SalesRecord.received_date, SalesRecord.returns).where(SalesRecord.created_date.between(start, end), SalesRecord.returns != 0)
    category_skus = _category_skus(db, category)
    if category_skus is not None:
        returns_query = returns_query.where(SalesRecord.sku.in_(category_skus))
    rows = db.execute(returns_query).all()
    daily = defaultdict(lambda: {"cancellations": 0.0, "refunds": 0.0})
    cancellation_totals = defaultdict(float)
    refund_totals = defaultdict(float)
    for created_date, sku, received_date, return_quantity in rows:
        quantity = float(return_quantity)
        # UZUM 口径：Возвраты != 0 后，以 Дата получения 是否为空分流；
        # 趋势日期始终归属 Дата создания。
        if received_date is None:
            daily[created_date]["cancellations"] += quantity
            cancellation_totals[sku] += quantity
        else:
            daily[created_date]["refunds"] += quantity
            refund_totals[sku] += quantity
    trend = [{"date": day, **values} for day, values in sorted(daily.items())]
    _, inventory = _latest_inventory(db, end)
    catalog = _sku_catalog(db)

    def top_skus(totals: dict[str, float], value_key: str) -> list[dict]:
        result = []
        for sku, value in sorted(totals.items(), key=lambda item: item[1], reverse=True)[:10]:
            identity = inventory.get(sku) or catalog.get(sku, {})
            result.append({
                "sku": sku,
                "product": identity.get("product"),
                "seller_sku": identity.get("seller_sku"),
                "color": identity.get("color"),
                "memory": identity.get("memory"),
                "region": identity.get("region"),
                value_key: value,
            })
        return result

    return {
        "trend": trend,
        "top_cancellations": top_skus(cancellation_totals, "cancellations"),
        "top_refunds": top_skus(refund_totals, "refunds"),
    }

def _stocktake_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

def _stocktake_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, str):
        cleaned = value.replace(" ", "").replace("\u00a0", "").replace(",", ".")
        if cleaned in ("", "-", "—"):
            return 0.0
        return float(cleaned)
    return float(value)

def _header_col(sheet, labels: list[str], fallback: int) -> int:
    normalized_labels = [re.sub(r"\s+", "", label).lower() for label in labels]
    for column in range(1, sheet.max_column + 1):
        value = re.sub(r"\s+", "", _stocktake_text(sheet.cell(1, column).value)).lower()
        if any(label in value or value in label for label in normalized_labels if value):
            return column
    return fallback

def _stocktake_supply_col(sheet) -> int:
    detected = _header_col(sheet, ["实际供货", "供货量", "供货", "SI（P）", "SI(P)"], 0)
    if detected:
        return detected
    # 兼容两种供货底表：旧版 G 列=实际供货；新版新增 Shipping date 后 I 列=实际供货。
    for fallback in (9, 7):
        header = _stocktake_text(sheet.cell(1, fallback).value).lower()
        if "shipping" not in header and "purchase" not in header:
            return fallback
    return 9

def _mapping_rows_from_sheet(sheet) -> tuple[dict[str, dict], dict[str, int]]:
    group_col = _header_col(sheet, ["品类划分"], 1)
    category_col = _header_col(sheet, ["分类"], 2)
    mapping_id_col = _header_col(sheet, ["小米ID", "ID"], 3)
    sku_col = _header_col(sheet, ["Штрихкод", "条形码", "barcode"], 4)
    market_name_col = _header_col(sheet, ["Market Name", "SKU", "seller_sku"], 5)
    spu_col = _header_col(sheet, ["SPU"], 6)
    color_col = _header_col(sheet, ["颜色", "color"], 7)
    memory_col = _header_col(sheet, ["内存", "ROM", "RAM"], 8)
    spec_col = _header_col(sheet, ["规格", "EU", "RU"], 9)
    mapping_by_id: dict[str, dict] = {}
    mapping_order: dict[str, int] = {}
    previous_group = ""
    previous_category = ""
    for row in range(2, sheet.max_row + 1):
        xiaomi_id = _stocktake_text(sheet.cell(row, mapping_id_col).value)
        if not xiaomi_id:
            continue
        group = _stocktake_text(sheet.cell(row, group_col).value) or previous_group
        category = _stocktake_text(sheet.cell(row, category_col).value) or previous_category
        previous_group = group or previous_group
        previous_category = category or previous_category
        if xiaomi_id not in mapping_by_id:
            mapping_by_id[xiaomi_id] = {
                "group": group,
                "category": category,
                "xiaomi_id": xiaomi_id,
                "sku": _stocktake_text(sheet.cell(row, sku_col).value),
                "market_name": _stocktake_text(sheet.cell(row, market_name_col).value),
                "spu": _stocktake_text(sheet.cell(row, spu_col).value),
                "color": _stocktake_text(sheet.cell(row, color_col).value),
                "memory": _stocktake_text(sheet.cell(row, memory_col).value),
                "region": _stocktake_text(sheet.cell(row, spec_col).value),
            }
            mapping_order[xiaomi_id] = row
    return mapping_by_id, mapping_order

def _mapping_catalog(db: Session) -> dict[str, dict]:
    rows = db.execute(select(ProductMapping).order_by(ProductMapping.updated_at.desc(), ProductMapping.id.desc())).scalars().all()
    result: dict[str, dict] = {}
    for row in rows:
        if row.xiaomi_id not in result:
            result[row.xiaomi_id] = {
                "group": row.group or "",
                "category": row.category or "",
                "xiaomi_id": row.xiaomi_id,
                "sku": row.sku or "",
                "market_name": row.market_name or "",
                "spu": row.spu or "",
                "color": row.color or "",
                "memory": row.memory or "",
                "region": row.region or "",
            }
    return result

def _upsert_product_mappings_from_content(content: bytes, db: Session) -> int:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="产品信息映射表无法读取，请上传 .xlsx 或 .xlsm 文件") from exc
    if not workbook.worksheets:
        raise HTTPException(status_code=422, detail="产品信息映射表为空")
    mapping_by_id, _ = _mapping_rows_from_sheet(workbook.worksheets[0])
    if not mapping_by_id:
        raise HTTPException(status_code=422, detail="未在产品信息映射表中读取到小米 ID、条形码和 SKU 属性")
    existing = {item.xiaomi_id: item for item in db.execute(select(ProductMapping).where(ProductMapping.xiaomi_id.in_(mapping_by_id.keys()))).scalars().all()}
    for xiaomi_id, item in mapping_by_id.items():
        target = existing.get(xiaomi_id)
        if target is None:
            target = ProductMapping(xiaomi_id=xiaomi_id)
            db.add(target)
        target.sku = item.get("sku") or target.sku
        target.market_name = item.get("market_name") or target.market_name
        target.group = item.get("group") or target.group
        target.category = item.get("category") or target.category
        target.spu = item.get("spu") or target.spu
        target.color = item.get("color") or target.color
        target.memory = item.get("memory") or target.memory
        target.region = item.get("region") or target.region
    db.commit()
    return len(mapping_by_id)

def _parse_stocktake_workbook(content: bytes, product_mapping: dict[str, dict] | None = None) -> tuple[dict[str, float], list[dict]]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="库存盘点文件无法读取，请上传 .xlsx 或 .xlsm 文件") from exc
    if not workbook.worksheets:
        raise HTTPException(status_code=422, detail="库存盘点文件为空")

    supply_sheet = workbook[workbook.sheetnames[0]]
    mapping_sheet = workbook["Sheet2"] if "Sheet2" in workbook.sheetnames else (workbook.worksheets[1] if len(workbook.worksheets) > 1 else None)
    id_col = _header_col(supply_sheet, ["小米ID", "ID"], 2)
    supply_col = _stocktake_supply_col(supply_sheet)
    supply_by_xiaomi_id: dict[str, float] = defaultdict(float)
    supply_order: dict[str, int] = {}
    for row in range(2, supply_sheet.max_row + 1):
        xiaomi_id = _stocktake_text(supply_sheet.cell(row, id_col).value)
        if not xiaomi_id:
            continue
        supply_order.setdefault(xiaomi_id, row)
        supply_by_xiaomi_id[xiaomi_id] += _stocktake_number(supply_sheet.cell(row, supply_col).value)
    if not supply_by_xiaomi_id:
        raise HTTPException(status_code=422, detail="未在供货底表中读取到小米 ID 与实际供货量，请确认 B 列为小米 ID、实际供货列标题为“实际供货”")

    mapping_by_id: dict[str, dict] = {}
    missing_mapping = 0
    if mapping_sheet is not None:
        mapping_by_id, _ = _mapping_rows_from_sheet(mapping_sheet)
    if product_mapping:
        mapping_by_id.update(product_mapping)

    rows = []
    for index, (xiaomi_id, supply) in enumerate(supply_by_xiaomi_id.items(), start=1):
        identity = mapping_by_id.get(xiaomi_id)
        if identity is None:
            missing_mapping += 1
            identity = {"xiaomi_id": xiaomi_id, "sku": "", "market_name": "", "group": "", "category": "", "spu": "", "color": "", "memory": "", "region": ""}
        rows.append({**identity, "supply": supply, "mapping_found": bool(mapping_by_id.get(xiaomi_id)), "_order": supply_order.get(xiaomi_id, index)})
    return dict(supply_by_xiaomi_id), sorted(rows, key=lambda row: row["_order"])

def _latest_inventory_overall(db: Session) -> tuple[date | None, dict[str, float]]:
    snapshot = db.scalar(select(func.max(InventoryHistory.download_date)))
    if snapshot is None:
        return None, {}
    rows = db.execute(
        select(InventoryHistory.sku, func.sum(InventoryHistory.inventory))
        .where(InventoryHistory.download_date == snapshot)
        .group_by(InventoryHistory.sku)
    ).all()
    return snapshot, {sku: float(value or 0) for sku, value in rows}

def _stocktake_return_events(db: Session) -> tuple[dict[str, list[dict]], dict[str, dict[str, float]]]:
    rows = db.execute(
        select(SalesRecord.sku, SalesRecord.created_date, SalesRecord.received_date, SalesRecord.returns, SalesRecord.id)
        .where(SalesRecord.returns != 0)
        .order_by(SalesRecord.created_date.desc(), SalesRecord.id.desc())
    ).all()
    events: dict[str, list[dict]] = defaultdict(list)
    totals: dict[str, dict[str, float]] = defaultdict(lambda: {"cancellations": 0.0, "refunds": 0.0})
    for sku, created_date, received_date, quantity, record_id in rows:
        qty = float(quantity or 0)
        kind = "取消" if received_date is None else "退货"
        event_date = created_date if received_date is None else received_date
        if kind == "取消":
            totals[sku]["cancellations"] += qty
        else:
            totals[sku]["refunds"] += qty
        events[sku].append({"date": event_date, "kind": kind, "quantity": qty, "id": record_id})
    for sku, sku_events in events.items():
        sku_events.sort(key=lambda item: (item["date"], item["id"]), reverse=True)
    return events, totals

def _stocktake_event_note(events: list[dict], needed: float) -> str:
    if needed <= 0:
        return ""
    remaining = ceil(needed)
    notes = []
    for event in events:
        if remaining <= 0:
            break
        qty = max(1, ceil(abs(float(event["quantity"] or 0))))
        used = min(remaining, qty)
        suffix = f"×{used}" if used > 1 else ""
        notes.append(f"{event['date']} {event['kind']}{suffix}")
        remaining -= used
    if remaining > 0:
        notes.append(f"仍缺 {remaining} 条取消/退货记录可对应")
    return "；".join(notes)

def _stocktake_result(content: bytes, db: Session, mapping_content: bytes | None = None) -> dict:
    mapping_saved_count = 0
    if mapping_content:
        mapping_saved_count = _upsert_product_mappings_from_content(mapping_content, db)
    _, rows = _parse_stocktake_workbook(content, _mapping_catalog(db))
    snapshot, inventory = _latest_inventory_overall(db)
    lifetime_rows = db.execute(
        select(SalesRecord.sku, func.sum(SalesRecord.quantity - SalesRecord.returns))
        .group_by(SalesRecord.sku)
    ).all()
    lifetime_sales = {sku: float(value or 0) for sku, value in lifetime_rows}
    events, return_totals = _stocktake_return_events(db)

    result_rows = []
    for row in rows:
        sku = row.get("sku") or ""
        current_inventory = inventory.get(sku, 0.0)
        lifetime_so = lifetime_sales.get(sku, 0.0)
        expected = current_inventory + lifetime_so
        diff = round(float(row["supply"]) - expected, 6)
        matched = abs(diff) < 1e-6
        totals = return_totals.get(sku, {"cancellations": 0.0, "refunds": 0.0})
        result_rows.append({
            "group": row.get("group") or "",
            "category": row.get("category") or "",
            "xiaomi_id": row.get("xiaomi_id") or "",
            "sku": sku,
            "market_name": row.get("market_name") or "",
            "spu": row.get("spu") or "",
            "color": row.get("color") or "",
            "memory": row.get("memory") or "",
            "region": row.get("region") or "",
            "mapping_found": bool(row.get("mapping_found")),
            "supply": float(row["supply"]),
            "lifetime_so": lifetime_so,
            "inventory": current_inventory,
            "cancellations": totals["cancellations"],
            "refunds": totals["refunds"],
            "matched": matched,
            "difference": 0.0 if matched else diff,
            "event_note": _stocktake_event_note(events.get(sku, []), diff) if diff > 0 else "",
        })
    matched_count = sum(row["matched"] for row in result_rows)
    positive_count = sum(row["difference"] > 0 for row in result_rows)
    negative_count = sum(row["difference"] < 0 for row in result_rows)
    missing_mapping_count = sum(not row["mapping_found"] for row in result_rows)
    return {
        "snapshot_date": snapshot,
        "inventory_date": snapshot - timedelta(days=1) if snapshot else None,
        "mapping_saved_count": mapping_saved_count,
        "summary": {
            "total": len(result_rows),
            "matched": matched_count,
            "mismatched": len(result_rows) - matched_count,
            "positive": positive_count,
            "negative": negative_count,
            "missing_mapping": missing_mapping_count,
            "total_supply": sum(row["supply"] for row in result_rows),
            "total_expected": sum(row["inventory"] + row["lifetime_so"] for row in result_rows),
            "total_difference": sum(row["difference"] for row in result_rows),
        },
        "items": result_rows,
    }

def _stocktake_workbook(result: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存盘点明细"
    headers = ["品类划分", "分类", "小米ID", "Штрихкод", "产品SKU", "SPU", "颜色", "内存", "规格", "SI（P）/供货量", "SO TTL（建店至今去退）", "Inventory（现有库存）", "取消订单数", "退货量", "进销存量是否对应", "进销存差额", "最近取消/退货记录"]
    sheet.append(headers)
    for row in result["items"]:
        sheet.append([
            row["group"], row["category"], row["xiaomi_id"], row["sku"], row["market_name"], row["spu"], row["color"], row["memory"], row["region"],
            row["supply"], row["lifetime_so"], row["inventory"], row["cancellations"], row["refunds"],
            "✅" if row["matched"] else "❌", row["difference"], row["event_note"],
        ])
    header_fill = PatternFill("solid", fgColor="F3EAFD")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in sheet.columns:
        letter = column[0].column_letter
        max_length = max(len(_stocktake_text(cell.value)) for cell in column)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 34)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

@app.post("/api/stocktake/analyze")
async def stocktake_analyze(file: UploadFile = File(...), mapping_file: UploadFile | None = File(None), db: Session = Depends(get_db)):
    content = await file.read()
    mapping_content = await mapping_file.read() if mapping_file else None
    return _stocktake_result(content, db, mapping_content)

@app.post("/api/stocktake/export")
async def stocktake_export(file: UploadFile = File(...), mapping_file: UploadFile | None = File(None), db: Session = Depends(get_db)):
    content = await file.read()
    mapping_content = await mapping_file.read() if mapping_file else None
    result = _stocktake_result(content, db, mapping_content)
    payload = _stocktake_workbook(result)
    filename = quote("库存盘点明细.xlsx")
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )

def _inventory_plan_data(db: Session, end: date, category: str = "all") -> dict:
    source_snapshot, inventory = _latest_inventory(db, end)
    start14 = end - timedelta(days=13); start30 = end - timedelta(days=29)
    sales14_rows = db.execute(select(SalesRecord.sku, func.sum(SalesRecord.quantity - SalesRecord.returns)).where(SalesRecord.created_date.between(start14, end)).group_by(SalesRecord.sku)).all()
    sales30_rows = db.execute(select(SalesRecord.sku, func.sum(SalesRecord.quantity - SalesRecord.returns)).where(SalesRecord.created_date.between(start30, end)).group_by(SalesRecord.sku)).all()
    lifetime_rows = db.execute(select(SalesRecord.sku, func.sum(SalesRecord.quantity - SalesRecord.returns)).where(SalesRecord.created_date <= end).group_by(SalesRecord.sku)).all()
    sales14 = {row[0]: float(row[1] or 0) for row in sales14_rows}; sales30 = {row[0]: float(row[1] or 0) for row in sales30_rows}; lifetime_sales = {row[0]: float(row[1] or 0) for row in lifetime_rows}
    items = []
    category_skus = _category_skus(db, category)
    for sku, stock in inventory.items():
        if category_skus is not None and sku not in category_skus:
            continue
        avg = sales14.get(sku, 0) / 14
        dos = safe_divide(stock["inventory"], avg) if avg > 0 else None
        replenish = max(0, ceil(avg * 28 - stock["inventory"])) if dos is not None else 0
        status = "slow" if dos is None else ("replenish" if dos < 28 else ("watch" if dos < 35 else "healthy"))
        items.append({"sku": sku, "product": stock["product"], "seller_sku": stock.get("seller_sku"), "color": stock.get("color"), "memory": stock.get("memory"), "region": stock.get("region"), "inventory": stock["inventory"], "lifetime_sales": lifetime_sales.get(sku, 0), "sales_14d": sales14.get(sku, 0), "average_sales_14d": avg, "sales_30d": sales30.get(sku, 0), "dos": dos, "replenishment": replenish, "status": status})
    counts = {status: sum(item["status"] == status for item in items) for status in ("healthy", "watch", "replenish", "slow")}
    return {"snapshot_date": end if source_snapshot else None, "source_snapshot_date": source_snapshot, "required_source_date": end + timedelta(days=1), "counts": counts, "items": sorted(items, key=lambda item: (item["replenishment"], item["inventory"]), reverse=True)}

@app.get("/api/inventory/psi")
def inventory_psi(end: date = Query(...), search: str | None = None, category: str = "all", db: Session = Depends(get_db)):
    result = _inventory_plan_data(db, end, category)
    items = result["items"]
    if search:
        needle = search.lower()
        items = [item for item in items if needle in item["sku"].lower() or needle in (item["seller_sku"] or "").lower() or needle in (item["product"] or "").lower()]
    result["items"] = items
    result["counts"] = {status: sum(item["status"] == status for item in items) for status in ("healthy", "watch", "replenish", "slow")}
    return result

@app.get("/api/inventory/history")
def inventory_daily_history(start: date = Query(...), end: date = Query(...), sku: str | None = None, category: str = "all", db: Session = Depends(get_db)):
    inventory_query = select(
        InventoryHistory.download_date,
        InventoryHistory.sku,
        func.max(InventoryHistory.product),
        func.max(InventoryHistory.seller_sku),
        func.max(InventoryHistory.color),
        func.max(InventoryHistory.memory),
        func.max(InventoryHistory.region),
        func.sum(InventoryHistory.inventory),
    ).where(InventoryHistory.download_date.between(start + timedelta(days=1), end + timedelta(days=1)))
    sales_query = select(
        SalesRecord.created_date,
        SalesRecord.sku,
        func.sum(SalesRecord.quantity - SalesRecord.returns),
    ).where(SalesRecord.created_date.between(start, end))
    if sku:
        inventory_query = inventory_query.where(InventoryHistory.sku == sku)
        sales_query = sales_query.where(SalesRecord.sku == sku)
    category_skus = _category_skus(db, category)
    if category_skus is not None:
        inventory_query = inventory_query.where(InventoryHistory.sku.in_(category_skus))
        sales_query = sales_query.where(SalesRecord.sku.in_(category_skus))
    inventory_rows = db.execute(inventory_query.group_by(InventoryHistory.download_date, InventoryHistory.sku).order_by(InventoryHistory.download_date, InventoryHistory.sku)).all()
    sales_rows = db.execute(sales_query.group_by(SalesRecord.created_date, SalesRecord.sku)).all()
    sales_by_key = {(row[0], row[1]): float(row[2] or 0) for row in sales_rows}
    sales_by_day = defaultdict(float)
    for (day, _), value in sales_by_key.items():
        sales_by_day[day] += value

    daily = defaultdict(lambda: {"inventory": 0.0, "inbound": 0.0, "has_previous": False})
    previous_by_sku: dict[str, tuple[date, float]] = {}
    identities: dict[str, dict] = {}
    for source_day, barcode, product, seller_sku, color, memory, region, inventory in inventory_rows:
        day = source_day - timedelta(days=1)
        stock = float(inventory or 0)
        identities[barcode] = {"sku": barcode, "product": product, "seller_sku": seller_sku, "color": color, "memory": memory, "region": region}
        daily[day]["inventory"] += stock
        previous = previous_by_sku.get(barcode)
        if previous is not None:
            previous_day, previous_stock = previous
            interval_sales = sum(value for (sale_day, sale_sku), value in sales_by_key.items() if sale_sku == barcode and previous_day < sale_day <= day)
            daily[day]["has_previous"] = True
            daily[day]["inbound"] += max(0.0, stock - previous_stock + interval_sales)
        previous_by_sku[barcode] = (day, stock)

    ordered_days = sorted(set(daily) | set(sales_by_day))
    series = []
    previous_total: float | None = None
    for day in ordered_days:
        values = daily.get(day)
        current_total = values["inventory"] if values is not None else None
        series.append({
            "date": day,
            "inventory": current_total,
            "change": None if previous_total is None or current_total is None else current_total - previous_total,
            "sales": sales_by_day.get(day, 0.0),
            "inbound": values["inbound"] if values is not None and values["has_previous"] else None,
        })
        if current_total is not None:
            previous_total = current_total
    latest_inventory = _sku_catalog(db)
    catalog = [{"sku": barcode, "product": item.get("product"), "seller_sku": item.get("seller_sku"), "color": item.get("color"), "memory": item.get("memory"), "region": item.get("region")} for barcode, item in latest_inventory.items() if category_skus is None or barcode in category_skus]
    return {
        "formula": "库存表 DATE 先减 1 天匹配销售日；快照间推算入库 = max(0, 本次库存 - 上次库存 + 两次库存日间累计SO)",
        "selected_sku": sku,
        "series": series,
        "skus": sorted(catalog or identities.values(), key=lambda item: (item.get("seller_sku") or item["sku"])),
    }

def _period_metrics(db: Session, start: date, end: date, category: str = "all") -> dict:
    category_skus = _category_skus(db, category)
    sales_query = select(func.coalesce(func.sum(SalesRecord.quantity), 0), func.coalesce(func.sum(SalesRecord.returns), 0), func.coalesce(func.sum(SalesRecord.gmv), 0)).where(SalesRecord.created_date.between(start, end))
    traffic_query = select(func.sum(TrafficHistory.uv)).where(TrafficHistory.record_date.between(start, end))
    if category_skus is not None:
        sales_query = sales_query.where(SalesRecord.sku.in_(category_skus))
        traffic_query = traffic_query.where(TrafficHistory.sku.in_(category_skus))
    sales = db.execute(sales_query).one()
    uv = db.scalar(traffic_query)
    result = calculate_metrics(float(sales[0]), float(sales[1]), float(sales[2]), float(uv) if uv is not None else None)
    result["return_rate"] = safe_divide(float(sales[1]), float(sales[0]))
    sales_days_query = select(SalesRecord.created_date).where(SalesRecord.created_date.between(start, end)).distinct()
    traffic_days_query = select(TrafficHistory.record_date).where(TrafficHistory.record_date.between(start, end)).distinct()
    if category_skus is not None:
        sales_days_query = sales_days_query.where(SalesRecord.sku.in_(category_skus))
        traffic_days_query = traffic_days_query.where(TrafficHistory.sku.in_(category_skus))
    sales_days = set(db.scalars(sales_days_query))
    traffic_days = set(db.scalars(traffic_days_query))
    result["traffic_missing"] = not traffic_days or not sales_days.issubset(traffic_days)
    if uv is not None and traffic_days:
        matched_query = (
            select(func.coalesce(func.sum(SalesRecord.quantity), 0))
            .join(TrafficHistory, and_(SalesRecord.created_date == TrafficHistory.record_date, SalesRecord.sku == TrafficHistory.sku))
            .where(SalesRecord.created_date.between(start, end))
        )
        if category_skus is not None:
            matched_query = matched_query.where(SalesRecord.sku.in_(category_skus))
        matched_quantity = db.scalar(matched_query)
        result["cvr"] = safe_divide(float(matched_quantity or 0), float(uv))
    return result

def _period_diagnosis(db: Session, start: date, end: date, previous_start: date, previous_end: date, category: str = "all") -> dict:
    """Compare two equal-length periods using the Agent's fixed eight-step diagnostic order."""
    current = _period_metrics(db, start, end, category)
    previous = _period_metrics(db, previous_start, previous_end, category)
    changes = {key: _change(current.get(key), previous.get(key)) for key in ("so", "orders", "gmv", "asp", "uv", "cvr", "return_rate")}

    current_products = {item["sku"]: item for item in _sku_metrics(db, start, end, category)}
    previous_products = {item["sku"]: item for item in _sku_metrics(db, previous_start, previous_end, category)}
    catalog = _sku_catalog(db)
    is_growth = current["so"] > previous["so"]
    sku_rows = []
    for sku in set(current_products) | set(previous_products):
        current_item = current_products.get(sku, {})
        previous_item = previous_products.get(sku, {})
        identity = current_item or previous_item or catalog.get(sku, {})
        current_so = float(current_item.get("so", 0))
        previous_so = float(previous_item.get("so", 0))
        sku_rows.append({
            "sku": sku, "product": identity.get("product"), "seller_sku": identity.get("seller_sku"),
            "color": identity.get("color"), "memory": identity.get("memory"), "region": identity.get("region"),
            "current_so": current_so, "previous_so": previous_so, "delta_so": current_so - previous_so,
        })
    sku_drivers = sorted(sku_rows, key=lambda item: item["delta_so"], reverse=is_growth)[:5]

    current_categories = {item["category"]: item for item in _category_breakdown(list(current_products.values()))}
    previous_categories = {item["category"]: item for item in _category_breakdown(list(previous_products.values()))}
    category_drivers = []
    for label in set(current_categories) | set(previous_categories):
        current_so = float(current_categories.get(label, {}).get("so", 0))
        previous_so = float(previous_categories.get(label, {}).get("so", 0))
        category_drivers.append({"category": label, "current_so": current_so, "previous_so": previous_so, "delta_so": current_so - previous_so})
    category_drivers.sort(key=lambda item: item["delta_so"], reverse=is_growth)

    current_source, current_inventory = _latest_inventory(db, end)
    previous_source, _ = _latest_inventory(db, previous_end)
    stockouts = []
    for sku, previous_item in previous_products.items():
        stock = current_inventory.get(sku)
        if stock is not None and stock["inventory"] <= 0 and previous_item["so"] > 0:
            identity = stock or catalog.get(sku, {})
            stockouts.append({
                "sku": sku, "product": identity.get("product"), "seller_sku": identity.get("seller_sku"),
                "color": identity.get("color"), "memory": identity.get("memory"), "region": identity.get("region"),
                "previous_so": previous_item["so"], "current_so": current_products.get(sku, {}).get("so", 0),
                "inventory": stock["inventory"],
            })
    stockouts.sort(key=lambda item: item["previous_so"], reverse=True)

    category_skus = _category_skus(db, category)
    def available_days(model, field, period_start: date, period_end: date) -> set[date]:
        statement = select(field).where(field.between(period_start, period_end)).distinct()
        if category_skus is not None:
            statement = statement.where(model.sku.in_(category_skus))
        return set(db.scalars(statement))

    current_sales_days = available_days(SalesRecord, SalesRecord.created_date, start, end)
    previous_sales_days = available_days(SalesRecord, SalesRecord.created_date, previous_start, previous_end)
    current_traffic_days = available_days(TrafficHistory, TrafficHistory.record_date, start, end)
    previous_traffic_days = available_days(TrafficHistory, TrafficHistory.record_date, previous_start, previous_end)
    quality = {
        "sales_current": bool(current_sales_days), "sales_previous": bool(previous_sales_days),
        "traffic_current": bool(current_traffic_days) and current_sales_days.issubset(current_traffic_days),
        "traffic_previous": bool(previous_traffic_days) and previous_sales_days.issubset(previous_traffic_days),
        "inventory_current": current_source is not None, "inventory_previous": previous_source is not None,
    }
    quality["complete"] = all(quality.values())

    def metric_check(key: str, label: str, adverse_when_up: bool = False) -> dict:
        change = changes[key]
        if current.get(key) is None or previous.get(key) is None:
            status = "missing"
        elif change is None or change == 0:
            status = "neutral"
        elif (change > 0) == adverse_when_up:
            status = "negative"
        else:
            status = "positive"
        return {"key": key, "label": label, "current": current.get(key), "previous": previous.get(key), "change": change, "status": status}

    checks = [
        metric_check("uv", "UV是否下降"), metric_check("cvr", "CVR是否下降"), metric_check("asp", "ASP是否变化"),
        {"key": "stockout", "label": "重点SKU是否缺货", "status": "missing" if current_source is None else ("negative" if stockouts else "positive"), "count": len(stockouts)},
        metric_check("return_rate", "取消/退货率是否上升", adverse_when_up=True),
        {"key": "sku", "label": "哪些SKU对上升贡献最大" if is_growth else "哪些SKU对下降贡献最大", "status": "positive" if is_growth and any(item["delta_so"] > 0 for item in sku_drivers) else ("negative" if any(item["delta_so"] < 0 for item in sku_drivers) else "neutral"), "count": sum(item["delta_so"] > 0 if is_growth else item["delta_so"] < 0 for item in sku_drivers)},
        {"key": "category", "label": "哪些品类拉动最大" if is_growth else "哪些品类拖累最大", "status": "positive" if is_growth and any(item["delta_so"] > 0 for item in category_drivers) else ("negative" if any(item["delta_so"] < 0 for item in category_drivers) else "neutral"), "count": sum(item["delta_so"] > 0 if is_growth else item["delta_so"] < 0 for item in category_drivers)},
        {"key": "quality", "label": "数据是否完整", "status": "positive" if quality["complete"] else "missing"},
    ]
    so_change = changes["so"]
    if not quality["sales_current"] or not quality["sales_previous"]:
        headline, direction = "销售数据不完整，暂不能判断环比销量变化原因", "missing"
    elif so_change is None:
        headline, direction = f"本期实际销量为 {current['so']:,.0f}，上一阶段无可比基数", "flat"
    elif so_change < 0:
        headline, direction = f"本期实际销量较上一阶段下降 {abs(so_change):.1%}", "decline"
    elif so_change > 0:
        headline, direction = f"本期实际销量较上一阶段增长 {so_change:.1%}", "growth"
    else:
        headline, direction = "本期实际销量与上一阶段持平", "flat"
    return {
        "date": end, "previous_date": previous_end, "current_start": start, "current_end": end,
        "previous_start": previous_start, "previous_end": previous_end, "direction": direction,
        "headline": headline, "current": current, "previous": previous, "changes": changes, "checks": checks,
        "stockouts": stockouts[:5], "sku_drivers": sku_drivers, "category_drivers": category_drivers,
        "data_quality": quality,
    }

@app.get("/api/reports/weekly")
def weekly_report(start: date = Query(...), end: date = Query(...), category: str = "all", db: Session = Depends(get_db)):
    days = (end - start).days + 1; previous_end = start - timedelta(days=1); previous_start = previous_end - timedelta(days=days - 1)
    current = _period_metrics(db, start, end, category); previous = _period_metrics(db, previous_start, previous_end, category)
    comparisons = {key: safe_divide(current[key] - previous[key], previous[key]) if current.get(key) is not None and previous.get(key) is not None else None for key in ("so", "orders", "gmv", "asp", "uv", "cvr", "return_rate")}
    products = _sku_metrics(db, start, end, category)
    planning = _inventory_plan_data(db, end, category)
    planning_by_sku = {item["sku"]: item for item in planning["items"]}
    for item in products:
        stock_plan = planning_by_sku.get(item["sku"])
        if stock_plan:
            item["inventory"] = stock_plan["inventory"]
            item["dos"] = stock_plan["dos"]
            item["average_sales_14d"] = stock_plan["average_sales_14d"]
            item["sales_30d"] = stock_plan["sales_30d"]
            item["stock_status"] = stock_plan["status"]
        else:
            item["dos"] = None
    hot = sorted(products, key=lambda item: item["gmv"], reverse=True)[:10]
    phones = [item for item in products if _business_category(item) == "手机"]
    tablets = [item for item in products if _business_category(item) == "平板"]
    mobile = phones + tablets
    wearable = [item for item in products if _business_category(item) == "可穿戴及其他"]
    slow_inventory = [item for item in planning["items"] if item["inventory"] > 0]
    slow_phones = sorted([item for item in slow_inventory if _business_category(item) == "手机"], key=lambda item: (item["sales_30d"], -item["inventory"]))[:10]
    slow_tablets = sorted([item for item in slow_inventory if _business_category(item) == "平板"], key=lambda item: (item["sales_30d"], -item["inventory"]))[:10]
    slow = sorted(slow_inventory, key=lambda item: (item["sales_30d"], -item["inventory"]))[:10]
    categories = _category_breakdown(products)
    insights = []
    so_change = comparisons["so"]
    if so_change is not None:
        leading_sku = (hot[0].get("seller_sku") or hot[0].get("product") or hot[0]["sku"]) if hot else "暂无 SKU"
        insights.append(f"SO较上一周期{'增长' if so_change >= 0 else '下降'}{abs(so_change):.1%}，主要销售贡献来自 {leading_sku}。")
    if comparisons["asp"] is not None: insights.append(f"ASP较上一周期{'提升' if comparisons['asp'] >= 0 else '下降'}{abs(comparisons['asp']):.1%}，影响GMV与销量增幅的差异。")
    if comparisons["return_rate"] is not None:
        insights.append(f"取消/退款率为{current['return_rate']:.1%}，较上一周期{'上升' if comparisons['return_rate'] >= 0 else '下降'}{abs(comparisons['return_rate']):.1%}，{'对实际销量形成压力' if comparisons['return_rate'] > 0 else '退货取消影响有所减弱'}。")
    if current["traffic_missing"]: insights.append("所选周期流量数据不完整，UV/CVR变化原因暂不下结论。")
    start_source, start_inventory = _latest_inventory(db, start)
    end_source, end_inventory = _latest_inventory(db, end)
    if start_source and end_source:
        category_skus = _category_skus(db, category)
        if category_skus is not None:
            start_inventory = {sku: item for sku, item in start_inventory.items() if sku in category_skus}
            end_inventory = {sku: item for sku, item in end_inventory.items() if sku in category_skus}
        start_stock = sum(item["inventory"] for item in start_inventory.values()); end_stock = sum(item["inventory"] for item in end_inventory.values())
        stock_change = end_stock - start_stock
        stock_pct = safe_divide(stock_change, start_stock)
        pct_text = f"（{abs(stock_pct):.1%}）" if stock_pct is not None else ""
        insights.append(f"所选周期在售库存从{start_stock:,.0f}变为{end_stock:,.0f}，{'增加' if stock_change >= 0 else '减少'}{abs(stock_change):,.0f}{pct_text}；库存变化已纳入销量原因判断。")
    else:
        missing = [str(day + timedelta(days=1)) for day, source in ((start, start_source), (end, end_source)) if source is None]
        insights.append(f"库存表缺少 DATE={ '、'.join(missing) } 的精确数据，库存变化不做推测。")
    diagnosis = _period_diagnosis(db, start, end, previous_start, previous_end, category)
    return {"period": {"start": start, "end": end, "previous_start": previous_start, "previous_end": previous_end}, "current": current, "previous": previous, "comparisons": comparisons, "diagnosis": diagnosis, "hot_top10": hot, "phone_top10": sorted(phones, key=lambda item: item["gmv"], reverse=True)[:10], "tablet_top10": sorted(tablets, key=lambda item: item["gmv"], reverse=True)[:10], "mobile_top10": sorted(mobile, key=lambda item: item["gmv"], reverse=True)[:10], "wearable_top10": sorted(wearable, key=lambda item: item["gmv"], reverse=True)[:10], "slow_top10": slow, "slow_phone_top10": slow_phones, "slow_tablet_top10": slow_tablets, "categories": categories, "insights": insights}

@app.post("/api/uploads")
async def upload_data(
    sales: UploadFile = File(...),
    inventory: UploadFile | None = File(None),
    traffic: UploadFile | None = File(None),
    inventory_date: date | None = Form(None),
    db: Session = Depends(get_db),
):
    if PUBLIC_READ_ONLY:
        raise HTTPException(status_code=403, detail="共享访问地址为只读模式，请在店主电脑的本机地址上传数据")
    results = [await import_file(db, "sales", sales, skip_duplicate=True)]
    if inventory:
        results.append(await import_file(db, "inventory", inventory, inventory_date, skip_duplicate=True))
    if traffic:
        results.append(await import_file(db, "traffic", traffic, skip_duplicate=True))
    return {"message": "数据上传并入库成功", "files": results}

# The production build is also served by FastAPI so the public tunnel needs a
# single origin. API routes must be registered before this catch-all mount.
FRONTEND_DIST = Path(__file__).resolve().parents[2] / "frontend" / "dist"
if FRONTEND_DIST.exists():
    app.mount("/", StaticFiles(directory=FRONTEND_DIST, html=True), name="frontend")
